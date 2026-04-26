from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="DCEBFF")
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7E2"),
    right=Side(style="thin", color="D0D7E2"),
    top=Side(style="thin", color="D0D7E2"),
    bottom=Side(style="thin", color="D0D7E2"),
)
MXN_FORMAT = '$#,##0.00'
DECIMAL_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00%'


def _apply_header_style(row) -> None:
    for cell in row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER


def _apply_borders(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                cell.border = THIN_BORDER


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 42)


def _write_rows(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    _apply_header_style(ws[1])
    for row in rows:
        ws.append(row)
    _apply_borders(ws)
    _auto_width(ws)


def _build_rr1_sheet(ws_rr1, reports: dict[str, object]) -> None:
    rr1_headers = [
        "UUID",
        "RFC emisor",
        "Proveedor",
        "Fecha",
        "Moneda",
        "Total original",
        "Total MXN",
        "Estatus SAT",
        "Riesgo",
        "Motivo",
    ]
    rr1_rows = [
        [
            row.get("uuid"),
            row.get("rfc_emisor"),
            row.get("proveedor"),
            row.get("fecha"),
            row.get("moneda"),
            row.get("total_original"),
            row.get("total_mxn"),
            row.get("estatus_sat"),
            row.get("riesgo"),
            row.get("motivo"),
        ]
        for row in reports.get("rr1", [])
    ]
    _write_rows(ws_rr1, rr1_headers, rr1_rows)
    for cell in ws_rr1["F"][1:]:
        cell.number_format = DECIMAL_FORMAT
    for cell in ws_rr1["G"][1:]:
        cell.number_format = MXN_FORMAT


def _build_rr9_sheet(ws_rr9, reports: dict[str, object]) -> None:
    rr9_headers = [
        "RFC emisor",
        "Proveedor",
        "Total MXN",
        "Facturas",
        "Canceladas",
        "Score riesgo",
        "Monedas usadas",
        "Operaciones repetidas",
        "Riesgo acumulado",
        "Requiere contrato",
        "Motivo",
    ]
    rr9_rows = [
        [
            row.get("rfc_emisor"),
            row.get("proveedor"),
            row.get("total_mxn"),
            row.get("facturas"),
            row.get("canceladas"),
            row.get("score_riesgo"),
            row.get("monedas_usadas"),
            row.get("operaciones_repetidas"),
            row.get("riesgo_acumulado"),
            "SI" if row.get("flag_requiere_contrato") else "NO",
            row.get("motivo"),
        ]
        for row in reports.get("rr9", [])
    ]
    _write_rows(ws_rr9, rr9_headers, rr9_rows)
    for cell in ws_rr9["C"][1:]:
        cell.number_format = MXN_FORMAT


def _build_resumen_riesgos_sheet(ws_resumen_riesgos, reports: dict[str, object]) -> None:
    headers = ["Indicador", "Valor", "Detalle"]
    rows = [
        [row.get("indicador"), row.get("valor"), row.get("detalle")]
        for row in reports.get("resumen_riesgos", [])
    ]
    _write_rows(ws_resumen_riesgos, headers, rows)


def _build_conciliacion_sheet(ws_conciliacion, reconciliation_rows: list[dict[str, object]]) -> None:
    headers = [
        "Fecha movimiento",
        "Descripcion",
        "Referencia",
        "Cargo",
        "Abono",
        "Monto",
        "Origen",
        "Estado conciliacion",
        "Score",
        "Motivo",
        "UUID CFDI relacionado",
        "Proveedor CFDI",
        "Total CFDI MXN",
    ]
    rows = [
        [
            row.get("fecha"),
            row.get("descripcion"),
            row.get("referencia"),
            row.get("cargo"),
            row.get("abono"),
            row.get("monto"),
            row.get("origen"),
            row.get("match_status"),
            row.get("match_score"),
            row.get("match_reason"),
            row.get("matched_invoice_uuid"),
            row.get("matched_invoice_provider"),
            row.get("matched_invoice_total_mxn"),
        ]
        for row in reconciliation_rows
    ]
    _write_rows(ws_conciliacion, headers, rows)
    for cell in ws_conciliacion["D"][1:] + ws_conciliacion["E"][1:] + ws_conciliacion["F"][1:] + ws_conciliacion["I"][1:]:
        cell.number_format = DECIMAL_FORMAT
    for cell in ws_conciliacion["M"][1:]:
        cell.number_format = MXN_FORMAT


def generate_excel_report(
    reports_bundle: dict[str, object],
    report_mode: str = "full",
    reconciliation_rows: list[dict[str, object]] | None = None,
) -> bytes:
    summary = reports_bundle["summary"]
    reports = reports_bundle["reports"]

    wb = Workbook()
    if report_mode == "rr1":
        ws_rr1 = wb.active
        ws_rr1.title = "RR1"
        _build_rr1_sheet(ws_rr1, reports)
        ws_resumen_riesgos = wb.create_sheet("RESUMEN_RIESGOS")
        _build_resumen_riesgos_sheet(ws_resumen_riesgos, reports)
    elif report_mode == "rr9":
        ws_rr9 = wb.active
        ws_rr9.title = "RR9"
        _build_rr9_sheet(ws_rr9, reports)
        ws_resumen_riesgos = wb.create_sheet("RESUMEN_RIESGOS")
        _build_resumen_riesgos_sheet(ws_resumen_riesgos, reports)
    else:
        ws_resumen = wb.active
        ws_resumen.title = "RESUMEN"
        resumen_headers = ["Indicador", "Valor"]
        resumen_rows = [
            ["Total facturado", summary["total_facturado"]],
            ["Facturas", summary["facturas"]],
            ["Vigentes", summary["vigentes"]],
            ["Canceladas", summary["canceladas"]],
            ["Proveedores unicos", summary["proveedores_unicos"]],
            ["Sin validacion SAT", summary["sin_validacion_sat"]],
            ["Total IVA trasladado", summary["total_iva_trasladado"]],
            ["Total IVA retenido", summary["total_iva_retenido"]],
            ["Total ISR retenido", summary["total_isr_retenido"]],
            ["Total impuestos netos", summary["total_impuestos_netos"]],
            ["Aviso", "Esto es extraccion y resumen, no calculo fiscal profesional"],
        ]
        _write_rows(ws_resumen, resumen_headers, resumen_rows)
        for cell in ws_resumen["B"][1:]:
            if isinstance(cell.value, (int, float)) and ws_resumen[f"A{cell.row}"].value in {
                "Total facturado",
                "Total IVA trasladado",
                "Total IVA retenido",
                "Total ISR retenido",
                "Total impuestos netos",
            }:
                cell.number_format = MXN_FORMAT

        ws_control = wb.create_sheet("CONTROL")
        control_headers = [
            "UUID",
            "RFC emisor",
            "Nombre",
            "Subtotal",
            "Descuento",
            "Moneda original",
            "Total original",
            "Tipo de cambio usado",
            "Fuente tipo de cambio",
            "Total MXN",
            "IVA trasladado",
            "IVA retenido",
            "ISR retenido",
            "IEPS",
            "Total impuestos",
            "Estatus SAT",
            "Riesgo",
        ]
        control_rows = [
            [
                row.get("uuid"),
                row.get("rfc_emisor"),
                row.get("razon_social"),
                row.get("subtotal"),
                row.get("descuento"),
                row.get("moneda_original") or row.get("moneda"),
                row.get("total_original"),
                row.get("tipo_cambio_usado"),
                row.get("fuente_tipo_cambio"),
                row.get("total_mxn"),
                row.get("iva_trasladado"),
                row.get("iva_retenido"),
                row.get("isr_retenido"),
                row.get("ieps_trasladado"),
                row.get("total_impuestos"),
                row.get("estatus_sat"),
                row.get("riesgo"),
            ]
            for row in reports["control"]
        ]
        _write_rows(ws_control, control_headers, control_rows)
        for cell in (
            ws_control["D"][1:]
            + ws_control["E"][1:]
            + ws_control["G"][1:]
            + ws_control["K"][1:]
            + ws_control["L"][1:]
            + ws_control["M"][1:]
            + ws_control["N"][1:]
            + ws_control["O"][1:]
        ):
            cell.number_format = DECIMAL_FORMAT
        for cell in ws_control["J"][1:]:
            cell.number_format = MXN_FORMAT

        ws_proveedores = wb.create_sheet("PROVEEDORES")
        proveedores_headers = [
            "RFC",
            "Nombre",
            "Total facturado MXN",
            "Facturas",
            "Canceladas",
            "% canceladas",
            "IVA total",
            "Score",
            "Nivel",
        ]
        proveedores_rows = [
            [
                row.get("rfc_emisor"),
                row.get("razon_social"),
                row.get("total_facturado"),
                row.get("facturas"),
                row.get("canceladas"),
                (row.get("porcentaje_canceladas", 0) or 0) / 100,
                row.get("iva_total"),
                row.get("score_riesgo"),
                row.get("nivel_riesgo"),
            ]
            for row in reports["proveedores"]
        ]
        _write_rows(ws_proveedores, proveedores_headers, proveedores_rows)
        for cell in ws_proveedores["C"][1:] + ws_proveedores["G"][1:]:
            cell.number_format = MXN_FORMAT
        for cell in ws_proveedores["F"][1:]:
            cell.number_format = PERCENT_FORMAT

        ws_riesgos = wb.create_sheet("RIESGOS")
        riesgos_headers = ["UUID", "Moneda original", "Total original", "Total MXN", "Tipo riesgo", "Nivel", "Detalle"]
        riesgos_rows = [
            [
                row.get("uuid"),
                row.get("moneda_original"),
                row.get("total_original"),
                row.get("total_mxn"),
                row.get("detalle_riesgo") or row.get("riesgo") or "-",
                row.get("riesgo"),
                row.get("detalle_riesgo") or "-",
            ]
            for row in reports["riesgos"]
        ]
        _write_rows(ws_riesgos, riesgos_headers, riesgos_rows)
        for cell in ws_riesgos["C"][1:]:
            cell.number_format = DECIMAL_FORMAT
        for cell in ws_riesgos["D"][1:]:
            cell.number_format = MXN_FORMAT

        ws_rr1 = wb.create_sheet("RR1")
        _build_rr1_sheet(ws_rr1, reports)

        ws_rr9 = wb.create_sheet("RR9")
        _build_rr9_sheet(ws_rr9, reports)

        ws_resumen_riesgos = wb.create_sheet("RESUMEN_RIESGOS")
        _build_resumen_riesgos_sheet(ws_resumen_riesgos, reports)

        if reconciliation_rows is not None:
            ws_conciliacion = wb.create_sheet("CONCILIACION")
            _build_conciliacion_sheet(ws_conciliacion, reconciliation_rows)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
