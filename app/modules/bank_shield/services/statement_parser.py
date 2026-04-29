from __future__ import annotations

"""Bank Shield v0.1 statement parser service.

Current implementation is migrated into the Bank Shield module while legacy
imports remain supported through temporary passthrough adapters.
"""

import csv
import hashlib
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

from openpyxl import load_workbook


HEADER_ALIASES = {
    "fecha": {
        "fecha",
        "fecha_operacion",
        "fecha operación",
        "fecha de operación",
        "fecha_movimiento",
        "fecha movimiento",
    },
    "descripcion": {
        "descripcion",
        "descripción",
        "descripcion movimiento",
        "concepto",
        "concepto movimiento",
        "detalle",
        "movimiento",
    },
    "referencia": {"referencia", "ref", "folio", "uuid", "referencia bancaria"},
    "cargo": {"cargo", "retiro", "egreso", "debito", "débito"},
    "abono": {"abono", "deposito", "depósito", "ingreso", "credito", "crédito"},
    "monto": {"monto", "importe", "total", "amount", "deposito", "depósito", "cargo", "abono"},
    "moneda": {"moneda", "currency", "divisa"},
}
DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d")


@dataclass(slots=True)
class ParsedBankTransaction:
    fecha: str | None
    descripcion: str
    referencia: str | None
    cargo: float
    abono: float
    monto: float
    tipo_movimiento: str
    moneda: str
    raw_hash: str


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _resolve_header_map(headers: list[str]) -> dict[str, int]:
    resolved: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        for canonical, aliases in HEADER_ALIASES.items():
            normalized_aliases = {_normalize_header(alias) for alias in aliases}
            if normalized in normalized_aliases and canonical not in resolved:
                resolved[canonical] = index
    return resolved


def _coerce_decimal(value: object) -> Decimal:
    text = str(value or "").strip()
    if not text:
        return Decimal("0")
    text = text.replace("$", "").replace(",", "")
    try:
        return Decimal(text)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def _normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _normalize_date(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(text, date_format).date().isoformat()
        except ValueError:
            continue
    if len(text) >= 10:
        return text[:10]
    return text or None


def _build_raw_hash(
    *,
    fecha: str | None,
    descripcion: str,
    referencia: str | None,
    cargo: float,
    abono: float,
    monto: float,
) -> str:
    signature = "|".join(
        [
            fecha or "",
            descripcion.upper(),
            str(referencia or "").upper(),
            f"{cargo:.2f}",
            f"{abono:.2f}",
            f"{monto:.2f}",
        ]
    )
    return hashlib.sha256(signature.encode("utf-8")).hexdigest()


def _build_transaction(row_values: dict[str, object]) -> ParsedBankTransaction | None:
    fecha = _normalize_date(row_values.get("fecha"))
    descripcion = _normalize_text(row_values.get("descripcion"))
    referencia = _normalize_text(row_values.get("referencia")) or None
    moneda = (_normalize_text(row_values.get("moneda")) or "MXN").upper()

    cargo_decimal = _coerce_decimal(row_values.get("cargo"))
    abono_decimal = _coerce_decimal(row_values.get("abono"))
    monto_decimal = _coerce_decimal(row_values.get("monto"))

    if monto_decimal == 0:
        if abono_decimal != 0:
            monto_decimal = abs(abono_decimal)
        elif cargo_decimal != 0:
            monto_decimal = abs(cargo_decimal)
    elif cargo_decimal == 0 and abono_decimal == 0:
        if monto_decimal < 0:
            cargo_decimal = -abs(monto_decimal)
        else:
            abono_decimal = abs(monto_decimal)

    if cargo_decimal > 0:
        cargo_decimal = -cargo_decimal
    if abono_decimal < 0:
        abono_decimal = abs(abono_decimal)

    monto_decimal = abs(monto_decimal)
    if monto_decimal == 0 or not descripcion:
        return None

    cargo = float(cargo_decimal)
    abono = float(abono_decimal)
    monto = float(monto_decimal)
    tipo_movimiento = "ABONO" if abono > 0 else "CARGO"
    return ParsedBankTransaction(
        fecha=fecha,
        descripcion=descripcion,
        referencia=referencia,
        cargo=cargo,
        abono=abono,
        monto=monto,
        tipo_movimiento=tipo_movimiento,
        moneda=moneda,
        raw_hash=_build_raw_hash(
            fecha=fecha,
            descripcion=descripcion,
            referencia=referencia,
            cargo=cargo,
            abono=abono,
            monto=monto,
        ),
    )


def _parse_rows(rows: list[list[object]]) -> list[ParsedBankTransaction]:
    if not rows:
        return []
    headers = [str(value or "") for value in rows[0]]
    header_map = _resolve_header_map(headers)
    if "fecha" not in header_map or "descripcion" not in header_map:
        raise ValueError("No fue posible identificar las columnas fecha y descripcion del estado bancario.")
    if not any(key in header_map for key in ("cargo", "abono", "monto")):
        raise ValueError("No fue posible identificar columnas de importe en el estado bancario.")

    transactions: list[ParsedBankTransaction] = []
    for raw_row in rows[1:]:
        row_values = {
            canonical: raw_row[index] if index < len(raw_row) else None
            for canonical, index in header_map.items()
        }
        transaction = _build_transaction(row_values)
        if transaction is not None:
            transactions.append(transaction)
    return transactions


def _parse_csv(file_bytes: bytes) -> list[ParsedBankTransaction]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(StringIO(text))
    rows = [row for row in reader if any(str(cell or "").strip() for cell in row)]
    return _parse_rows(rows)


def _parse_xlsx(file_bytes: bytes) -> list[ParsedBankTransaction]:
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = [list(row) for row in worksheet.iter_rows(values_only=True) if any(str(cell or "").strip() for cell in row)]
    return _parse_rows(rows)


def parse_bank_statement(file_bytes: bytes, filename: str) -> list[ParsedBankTransaction]:
    normalized_name = str(filename or "").strip().lower()
    if normalized_name.endswith(".csv"):
        return _parse_csv(file_bytes)
    if normalized_name.endswith(".xlsx"):
        return _parse_xlsx(file_bytes)
    raise ValueError("Solo se aceptan estados bancarios CSV o XLSX.")
