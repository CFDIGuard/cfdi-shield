from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.db.base import Base
from app.db import init_db as init_db_module
from app.db import session as session_module
from app.main import app
from app.repositories.bank_transaction_repository import BankTransactionRepository
from app.models.invoice import Invoice
from app.repositories.invoice_repository import InvoiceRepository
from app.repositories.user_repository import UserRepository
from app.schemas.invoice import InvoiceCreate
from app.schemas.payment_complement import PaymentComplementProcessedData
from app.services.auth_service import create_session_token, hash_password
import app.web_deps as web_deps_module


def _make_invoice(user_id: int, uuid: str, total: float, razon_social: str) -> InvoiceCreate:
    return InvoiceCreate(
        user_id=user_id,
        uuid=uuid,
        archivo=f"{uuid}.xml",
        razon_social=razon_social,
        rfc_emisor="AAA010101AAA",
        rfc_receptor="BBB010101BBB",
        folio="F-1",
        fecha_emision="2026-04-25T10:00:00",
        mes="2026-04",
        subtotal=total,
        total=total,
        total_original=total,
        iva=0,
        iva_retenido=0,
        isr_retenido=0,
        moneda="MXN",
        moneda_original="MXN",
        tipo_cambio_xml=None,
        tipo_cambio_usado=1,
        total_mxn=total,
        fuente_tipo_cambio="MXN",
        fecha_tipo_cambio="2026-04-25",
        metodo_pago="PUE",
        estatus_sat="VIGENTE",
        riesgo="BAJO",
        score_proveedor=0,
        detalle_riesgo="",
        sat_validado_at=None,
    )


def test_multiuser_excel_delete_and_get_isolation(tmp_path, monkeypatch):
    db_path = tmp_path / "multiuser.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    testing_session_local = sessionmaker(bind=engine, autoflush=False, autocommit=False)

    monkeypatch.setattr(session_module, "engine", engine)
    monkeypatch.setattr(session_module, "SessionLocal", testing_session_local)
    monkeypatch.setattr(init_db_module, "engine", engine)
    monkeypatch.setattr(init_db_module, "_initialized", False)
    monkeypatch.setattr(web_deps_module, "SessionLocal", testing_session_local)

    Base.metadata.create_all(bind=engine)
    init_db_module.ensure_db_initialized()

    with testing_session_local() as db:
        user_repo = UserRepository(db)
        user_a = user_repo.create("a@example.com", hash_password("password123"))
        user_b = user_repo.create("b@example.com", hash_password("password123"))

        repo_a = InvoiceRepository(db, user_id=user_a.id)
        repo_b = InvoiceRepository(db, user_id=user_b.id)
        invoice_a_payload = _make_invoice(user_a.id, "AAAAAAAA-1111-4111-8111-AAAAAAAAAAAA", 100.0, "Proveedor A")
        invoice_a_payload.estatus_sat = "CANCELADO"
        invoice_a_payload.riesgo = "ALTO"
        invoice_a_payload.detalle_riesgo = "CFDI cancelado"
        invoice_a = repo_a.create(invoice_a_payload)

        invoice_a2_payload = _make_invoice(user_a.id, "AAAAAAAA-3333-4333-8333-AAAAAAAAAAAA", 150.0, "Proveedor A2")
        invoice_a2_payload.rfc_emisor = "CCC010101CCC"
        invoice_a2_payload.estatus_sat = "VIGENTE"
        invoice_a2_payload.riesgo = "BAJO"
        invoice_a2_payload.detalle_riesgo = ""
        repo_a.create(invoice_a2_payload)

        payment_a_payload = _make_invoice(user_a.id, "AAAAAAAA-4444-4444-8444-AAAAAAAAAAAA", 0.0, "Proveedor A")
        payment_a_payload.tipo_comprobante = "P"
        payment_a_payload.moneda = "XXX"
        payment_a_payload.moneda_original = "XXX"
        payment_a_payload.metodo_pago = "PPD"
        payment_a_payload.payment_complements = [
            PaymentComplementProcessedData(
                related_invoice_uuid=invoice_a.uuid,
                fecha_pago="2026-04-25T11:00:00",
                moneda_pago="MXN",
                monto_pago=100.0,
                parcialidad=1,
                saldo_anterior=100.0,
                importe_pagado=100.0,
                saldo_insoluto=0.0,
            )
        ]
        repo_a.create(payment_a_payload)
        refreshed_invoice_a = repo_a.get_by_id(invoice_a.id)
        assert refreshed_invoice_a is not None
        assert refreshed_invoice_a.estado_pago == "PAGADA"
        assert float(refreshed_invoice_a.total_pagado or 0) == 100.0
        assert float(refreshed_invoice_a.saldo_pendiente or 0) == 0.0

        invoice_b_payload = _make_invoice(user_b.id, "BBBBBBBB-2222-4222-8222-BBBBBBBBBBBB", 200.0, "Proveedor B")
        invoice_b_payload.estatus_sat = "SIN_VALIDACION"
        invoice_b_payload.riesgo = "MEDIO"
        invoice_b_payload.detalle_riesgo = "CFDI sin validacion SAT"
        invoice_b = repo_b.create(invoice_b_payload)

    client = TestClient(app)
    cookie_name = settings.session_cookie_name
    cookie_a = {cookie_name: create_session_token(user_a.id)}
    cookie_b = {cookie_name: create_session_token(user_b.id)}

    response_a = client.get("/api/v1/dashboard/export-excel", cookies=cookie_a)
    assert response_a.status_code == 200
    workbook_a = load_workbook(filename=BytesIO(response_a.content))
    control_a = workbook_a["CONTROL"]
    control_headers_a = [cell.value for cell in next(control_a.iter_rows(max_row=1))]
    assert "Total pagado" in control_headers_a
    assert "Saldo pendiente" in control_headers_a
    assert "Estado pago" in control_headers_a
    values_a = "\n".join("" if cell is None else str(cell) for row in control_a.iter_rows(values_only=True) for cell in row)
    assert "Proveedor A" in values_a
    assert "Proveedor B" not in values_a
    assert "BBBBBBBB-2222-4222-8222-BBBBBBBBBBBB" not in values_a
    resumen_a = workbook_a["RESUMEN"]
    resumen_values_a = {row[0]: row[1] for row in resumen_a.iter_rows(values_only=True) if row and row[0]}
    assert resumen_values_a.get("Facturas pagadas") == 1

    response_b = client.get("/api/v1/dashboard/export-excel", cookies=cookie_b)
    assert response_b.status_code == 200
    workbook_b = load_workbook(filename=BytesIO(response_b.content))
    control_b = workbook_b["CONTROL"]
    values_b = "\n".join("" if cell is None else str(cell) for row in control_b.iter_rows(values_only=True) for cell in row)
    assert "Proveedor B" in values_b
    assert "Proveedor A" not in values_b
    assert "AAAAAAAA-1111-4111-8111-AAAAAAAAAAAA" not in values_b

    rr1_a = client.get("/api/v1/dashboard/export-rr1-excel", cookies=cookie_a)
    assert rr1_a.status_code == 200
    rr1_book_a = load_workbook(filename=BytesIO(rr1_a.content))
    rr1_sheet_a = rr1_book_a["RR1"]
    rr1_values_a = "\n".join("" if cell is None else str(cell) for row in rr1_sheet_a.iter_rows(values_only=True) for cell in row)
    assert "Proveedor A" in rr1_values_a
    assert "Proveedor B" not in rr1_values_a

    rr9_b = client.get("/api/v1/dashboard/export-rr9-excel", cookies=cookie_b)
    assert rr9_b.status_code == 200
    rr9_book_b = load_workbook(filename=BytesIO(rr9_b.content))
    rr9_sheet_b = rr9_book_b["RR9"]
    rr9_values_b = "\n".join("" if cell is None else str(cell) for row in rr9_sheet_b.iter_rows(values_only=True) for cell in row)
    assert "Proveedor B" in rr9_values_b
    assert "Proveedor A" not in rr9_values_b

    filtered_excel_a = client.get(
        "/api/v1/dashboard/export-excel?rfc_emisor=AAA010101AAA&estatus_sat=CANCELADO",
        cookies=cookie_a,
    )
    assert filtered_excel_a.status_code == 200
    filtered_book_a = load_workbook(filename=BytesIO(filtered_excel_a.content))
    filtered_control_a = filtered_book_a["CONTROL"]
    filtered_values_a = "\n".join("" if cell is None else str(cell) for row in filtered_control_a.iter_rows(values_only=True) for cell in row)
    assert "Proveedor A" in filtered_values_a
    assert "Proveedor A2" not in filtered_values_a
    assert "Proveedor B" not in filtered_values_a

    rr1_filtered_page = client.get(
        "/reports/rr1?rfc_emisor=AAA010101AAA&estatus_sat=CANCELADO",
        cookies=cookie_a,
    )
    assert rr1_filtered_page.status_code == 200
    assert "Proveedor A" in rr1_filtered_page.text
    assert "Proveedor A2" not in rr1_filtered_page.text
    assert "Proveedor B" not in rr1_filtered_page.text

    rr9_filtered_page = client.get(
        "/reports/rr9?rfc_emisor=BBB010101BBB&riesgo=MEDIO",
        cookies=cookie_b,
    )
    assert rr9_filtered_page.status_code == 200
    assert "Proveedor B" in rr9_filtered_page.text
    assert "Proveedor A" not in rr9_filtered_page.text

    bank_csv = (
        "fecha,descripcion,referencia,monto\n"
        "2026-04-25,Pago Proveedor A AAA010101AAA,AAAAAAAA-1111-4111-8111-AAAAAAAAAAAA,100.00\n"
        "2026-04-25,Pago Proveedor A2,,151.00\n"
        "2026-04-25,Movimiento sin match,,999.00\n"
    ).encode("utf-8")
    upload_reconciliation_a = client.post(
        "/reconciliation/upload",
        cookies=cookie_a,
        files={"file": ("estado.csv", bank_csv, "text/csv")},
        follow_redirects=False,
    )
    assert upload_reconciliation_a.status_code == 303

    reconciliation_page_a = client.get("/reconciliation", cookies=cookie_a)
    assert reconciliation_page_a.status_code == 200
    assert "Pago Proveedor A AAA010101AAA" in reconciliation_page_a.text
    assert "Pago Proveedor A2" in reconciliation_page_a.text
    assert "Movimiento sin match" in reconciliation_page_a.text
    assert "CONCILIADO" in reconciliation_page_a.text
    assert "POSIBLE" in reconciliation_page_a.text
    assert "PENDIENTE" in reconciliation_page_a.text

    reconciliation_page_b = client.get("/reconciliation", cookies=cookie_b)
    assert reconciliation_page_b.status_code == 200
    assert "Pago Proveedor A AAA010101AAA" not in reconciliation_page_b.text
    assert "Movimiento sin match" not in reconciliation_page_b.text

    with testing_session_local() as db:
        bank_repo_a = BankTransactionRepository(db, user_id=user_a.id)
        txs_a = bank_repo_a.list_all()
        assert len(txs_a) == 3
        posible_tx = next(tx for tx in txs_a if abs(tx.monto - 151.0) < 0.01)
        pendiente_tx = next(tx for tx in txs_a if abs(tx.monto - 999.0) < 0.01)
        conciliado_tx = next(tx for tx in txs_a if abs(tx.monto - 100.0) < 0.01)
        assert posible_tx.match_status == "POSIBLE"
        assert conciliado_tx.match_status == "CONCILIADO"

    confirm_posible = client.post(f"/reconciliation/confirm/{posible_tx.id}", cookies=cookie_a)
    assert confirm_posible.status_code == 200
    assert confirm_posible.json()["transaction"]["match_status"] == "CONCILIADO"
    assert confirm_posible.json()["transaction"]["origen"] == "MANUAL"

    assign_pending = client.post(
        f"/reconciliation/assign/{pendiente_tx.id}?invoice_id={invoice_a.id}",
        cookies=cookie_a,
    )
    assert assign_pending.status_code == 200
    assert assign_pending.json()["transaction"]["match_status"] == "CONCILIADO"
    assert assign_pending.json()["transaction"]["matched_invoice_uuid"] == invoice_a.uuid
    assert assign_pending.json()["transaction"]["origen"] == "MANUAL"

    reject_tx = client.post(f"/reconciliation/reject/{conciliado_tx.id}", cookies=cookie_a)
    assert reject_tx.status_code == 200
    assert reject_tx.json()["transaction"]["match_status"] == "PENDIENTE"
    assert reject_tx.json()["transaction"]["matched_invoice_uuid"] is None

    assign_foreign_invoice = client.post(
        f"/reconciliation/assign/{pendiente_tx.id}?invoice_id={invoice_b.id}",
        cookies=cookie_a,
    )
    assert assign_foreign_invoice.status_code == 404

    filtered_pending = client.get("/reconciliation?estado=PENDIENTE", cookies=cookie_a)
    assert filtered_pending.status_code == 200
    assert "Movimiento sin match" not in filtered_pending.text
    assert "Pago Proveedor A AAA010101AAA" in filtered_pending.text

    filtered_manual = client.get("/reconciliation?origen=MANUAL", cookies=cookie_a)
    assert filtered_manual.status_code == 200
    assert "Pago Proveedor A AAA010101AAA" in filtered_manual.text
    assert "Pago Proveedor A2" in filtered_manual.text

    search_provider = client.get("/reconciliation?busqueda=Proveedor%20A2", cookies=cookie_a)
    assert search_provider.status_code == 200
    assert "Pago Proveedor A2" in search_provider.text

    reconciliation_excel_a = client.get("/api/v1/dashboard/export-excel", cookies=cookie_a)
    assert reconciliation_excel_a.status_code == 200
    reconciliation_book_a = load_workbook(filename=BytesIO(reconciliation_excel_a.content))
    assert "CONCILIACION" in reconciliation_book_a.sheetnames
    reconciliation_sheet_a = reconciliation_book_a["CONCILIACION"]
    reconciliation_values_a = "\n".join(
        "" if cell is None else str(cell)
        for row in reconciliation_sheet_a.iter_rows(values_only=True)
        for cell in row
    )
    assert "Pago Proveedor A AAA010101AAA" in reconciliation_values_a
    assert "Pago Proveedor A2" in reconciliation_values_a
    assert "Movimiento sin match" in reconciliation_values_a

    reconciliation_excel_b = client.get("/api/v1/dashboard/export-excel", cookies=cookie_b)
    assert reconciliation_excel_b.status_code == 200
    reconciliation_book_b = load_workbook(filename=BytesIO(reconciliation_excel_b.content))
    assert "CONCILIACION" in reconciliation_book_b.sheetnames
    reconciliation_sheet_b = reconciliation_book_b["CONCILIACION"]
    reconciliation_values_b = "\n".join(
        "" if cell is None else str(cell)
        for row in reconciliation_sheet_b.iter_rows(values_only=True)
        for cell in row
    )
    assert "Pago Proveedor A AAA010101AAA" not in reconciliation_values_b
    assert "Movimiento sin match" not in reconciliation_values_b

    complement_sheet_a = workbook_a["COMPLEMENTOS_PAGO"]
    complement_values_a = "\n".join(
        "" if cell is None else str(cell)
        for row in complement_sheet_a.iter_rows(values_only=True)
        for cell in row
    )
    assert "AAAAAAAA-4444-4444-8444-AAAAAAAAAAAA" in complement_values_a
    assert "AAAAAAAA-1111-4111-8111-AAAAAAAAAAAA" in complement_values_a
    assert "BBBBBBBB-2222-4222-8222-BBBBBBBBBBBB" not in complement_values_a

    complement_sheet_b = workbook_b["COMPLEMENTOS_PAGO"]
    complement_values_b = "\n".join(
        "" if cell is None else str(cell)
        for row in complement_sheet_b.iter_rows(values_only=True)
        for cell in row
    )
    assert "AAAAAAAA-4444-4444-8444-AAAAAAAAAAAA" not in complement_values_b

    filtered_reconciliation_export = client.get(
        "/reconciliation/export-excel?origen=MANUAL&busqueda=Proveedor%20A2",
        cookies=cookie_a,
    )
    assert filtered_reconciliation_export.status_code == 200
    filtered_reconciliation_book = load_workbook(filename=BytesIO(filtered_reconciliation_export.content))
    filtered_reconciliation_sheet = filtered_reconciliation_book["CONCILIACION"]
    filtered_reconciliation_values = "\n".join(
        "" if cell is None else str(cell)
        for row in filtered_reconciliation_sheet.iter_rows(values_only=True)
        for cell in row
    )
    assert "Pago Proveedor A2" in filtered_reconciliation_values
    assert "Movimiento sin match" not in filtered_reconciliation_values

    get_b_as_a = client.get(f"/api/v1/invoices/{invoice_b.id}", cookies=cookie_a)
    assert get_b_as_a.status_code == 404

    delete_b_as_a = client.post(
        f"/invoices/{invoice_b.id}/delete",
        cookies=cookie_a,
        follow_redirects=False,
    )
    assert delete_b_as_a.status_code == 303

    with testing_session_local() as db:
        remaining_b = InvoiceRepository(db, user_id=user_b.id).get_by_id(invoice_b.id)
        assert remaining_b is not None
